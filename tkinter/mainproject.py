from tkinter import *
from tkinter.ttk import Combobox
from PIL import ImageTk,Image
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import re

class Lab(Tk):
    def __init__(self):
        super().__init__()
        self.title('😊😊 WE CARE 😊😊')
        self.geometry('700x600')

        #FUNCTION TO CALL FOR SECOND TAB
        def calculate():
            win=Tk()
            win.geometry('740x500')
            win.title('calci')
            win.config(bg='light yellow')
             
             #FIRST FRAME OF DETAILS OF SECOND WINDOW
            f1=Frame(win,bg='light yellow')
            Label(f1,text="* DETAILS -----  ",font=("calibriheading 15 bold"),bg='light yellow',fg='red').grid(row=1,ipady=10)
            Label(f1,text=f"NAME = {name.get()}",font=("ARIALBLACK 10 bold"),bg='light yellow').grid(column=0,row=2,ipady=10,sticky=W)
            Label(f1,text=f"AGE = {age.get()}",font=("ARIALBLACK 10 bold"),bg='light yellow').grid(column=0,row=3,ipady=10,sticky=W)
            Label(f1,text=f"GENDER = {combo.get()}",font=("ARIALBLACK 10 bold"),bg='light yellow').grid(column=0,row=4,ipady=10,sticky=W)
            Label(f1,text=f"Address = {addr.get()}",font=("ARIALBLACK 10 bold"),bg='light yellow').grid(column=0,row=7,sticky=W,ipady=10)
            Label(f1,text=f"Contact No ={phone.get()}",font=("ARIALBLACK 10 bold"),bg='light yellow').grid(column=0,row=8,sticky=W,ipady=10) 
            f1.grid(column=0)
            
             #SECOND FRAME FOR LINE 
            f2=Frame(win,bg='light yellow',pady=10) 
            Label(f2,text='--------------------------------------------------------------------------------------------------',font=(
            'Algebra 10 bold'),bg='light yellow',fg="red").grid(row=8,rowspan=1,sticky=W)
            f2.grid(column=0)
            
             #THIRD FRAME FOR RESULT/STATUS OF CALCULATIONS
            f3=Frame(win,padx=10,pady=20,bg="light yellow")
            bmi=calculate_bmi(weig.get(),heig.get())
            str1=bmi_index(bmi) 
            Label(f3,text=f" * BMI = {bmi}   AND  STATUS  -- {str1} ",font='baskervilleoldface 13 bold',bg="light green",fg="red").grid(column=0,row=9,ipadx=5,sticky=W)
            oxy1=oxystatus(oxy.get())
            Label(f3,text=f'* {oxy1}',font='baskervilleoldface 13 bold',bg="white",fg="black").grid(column=0,row=10,ipadx=10,ipady=5,rowspan=1,sticky=W)
            val=bpstatus(b_p.get())
            Label(f3,text=f'* BLOOD PRESSURE STATUS  - {val}',font='baskervilleoldface 13 bold',bg="red",fg="yellow").grid(column=0,row=12,ipadx=10,ipady=5,rowspan=1,sticky=W)
            f3.grid(column=0)
             
             #FOURTH FRAME FOR EXIT BUTTONS 
            f4=LabelFrame(win,padx=10)
            Button(f4,text="EXIT",font='arialblack 10 bold',fg='red',command=quit).grid(column=1,ipadx=20,row=18)
            Button(f4,text="SHOW DATA",font='arialblack 10 bold',fg='red',command=showwin).grid(column=0,row=18,sticky=NW)
            f4.grid(column=0)
            Label(win,text='MADE BY- SATYENDRA AND SARTHAK V',font="arialblak 8 bold ",bg='skyblue',fg='red').place(x=390,y=490)#grid(row=19,rowspan=1)
            
            
            #COMMAND TO SAVE DATA IN EXCEL FIL CALLED "sinchan.xlsx"
            file=openpyxl.load_workbook('shinchan2.xlsx')
            sheet=file.active
            #sheet.cell(column=1, row=sheet.max_row+1,value="current date and time")
            sheet.cell(column=1, row=sheet.max_row+1,value=name.get())
            sheet.cell(column=2, row=sheet.max_row,value=age.get())
            sheet.cell(column=3, row=sheet.max_row,value=combo.get())
            sheet.cell(column=4, row=sheet.max_row,value=phone.get())
            sheet.cell(column=5, row=sheet.max_row,value=addr.get())
            sheet.cell(column=6, row=sheet.max_row,value=heig.get())
            sheet.cell(column=7, row=sheet.max_row,value=weig.get())
            sheet.cell(column=8, row=sheet.max_row,value=bmi)
            sheet.cell(column=9, row=sheet.max_row,value=str1)
            sheet.cell(column=10, row=sheet.max_row,value=b1p.get())
            sheet.cell(column=11, row=sheet.max_row,value=val)
            sheet.cell(column=12, row=sheet.max_row,value=oxyg.get())
            sheet.cell(column=13, row=sheet.max_row,value=oxy1)
    
            file.save('shinchan2.xlsx') 
        
        def showwin():
    
            # initalise the tkinter GUI
            root = Tk()

            root.geometry("800x550") # set the root dimensions
            root.pack_propagate(False) # tells the root to not let the widgets inside it determine its size.
            root.resizable(0, 0) # makes the root window fixed in size.
            root.config(bg='light blue')

            # Frame for TreeView
            frame1 =LabelFrame(root, text="Excel Data")
            frame1.place(height=400, width=800)

            # Frame for open file dialog
            file_frame =LabelFrame(root,text="Open File",bg="light yellow")
            file_frame.place(height=100, width=700, rely=0.80, relx=0.10)
            def Filedialog():
                """This Function will open the file explorer and assign the chosen file path to label_file"""
                filename = filedialog.askopenfilename(initialdir="/",
                                                    title="Select A File",
                                                    filetype=(("xlsx files", "*.xlsx"),("All Files", "*.*")))
                label_file["text"] = filename
                return None
            
            # Buttons
            button1 =Button(file_frame, text="Browse A File", command=lambda:Filedialog())
            button1.place(rely=0.65, relx=0.50)
            
            def Load_excel_data():
                """If the file selected is valid this will load the file into the Treeview"""
                file_path = label_file["text"]
                try:
                    excel_filename = r"{}".format(file_path)
                    if excel_filename[-4:] == ".csv":
                        df = pd.read_csv(excel_filename)
                    else:
                        df = pd.read_excel(excel_filename)

                except ValueError:
                    messagebox.showerror("Information", "The file you have chosen is invalid")
                    return None
                except FileNotFoundError:
                    messagebox.showerror("Information", f"No such file as {file_path}")
                    return None

                clear_data()
                tv1["column"] = list(df.columns)
                tv1["show"] = "headings"
                for column in tv1["columns"]:
                    tv1.heading(column, text=column) # let the column heading = column name

                df_rows = df.to_numpy().tolist() # turns the dataframe into a list of lists
                for row in df_rows:
                    tv1.insert("", "end", values=row) # inserts each list into the treeview. For parameters see https://docs.python.org/3/library/tkinter.ttk.html#tkinter.ttk.Treeview.insert
                return None


            def clear_data():
                tv1.delete(*tv1.get_children())
                return None

            button2 = Button(file_frame, text="Load File", command=lambda: Load_excel_data())
            button2.place(rely=0.65, relx=0.30)

            # The file/file path text
            label_file = ttk.Label(file_frame, text="No File Selected")
            label_file.place(rely=0, relx=0)


            ## Treeview Widget
            tv1 = ttk.Treeview(frame1)
            tv1.place(relheight=1, relwidth=1) # set the height and width of the widget to 100% of its container (frame1).

            treescrolly =Scrollbar(frame1, orient="vertical", command=tv1.yview) # command means update the yaxis view of the widget
            treescrollx =Scrollbar(frame1, orient="horizontal", command=tv1.xview) # command means update the xaxis view of the widget
            tv1.configure(xscrollcommand=treescrollx.set, yscrollcommand=treescrolly.set) # assign the scrollbars to the Treeview Widget
            treescrollx.pack(side="bottom", fill="x") # make the scrollbar fill the x axis of the Treeview widget
            treescrolly.pack(side="right", fill="y") # make the scrollbar fill the y axis of the Treeview widget

  
         
         # RESET the content of text entry box   
        def clear():
            name.delete(0, END)
            age.delete(0, END)
            phone.delete(0, END)
            addr.delete(0, END)
            heig.delete(0, END)
            weig.delete(0, END)
            combo.set('SELECT')
            b1p.delete(0,END)
            oxyg.delete(0,END)
            #address_field.delete(0, END)
    
        icon= Image.open("D:\my python\python code\Tklibrary\download.jpeg")
        icon=icon.resize((700,600),Image.ANTIALIAS)
        self.photo=ImageTk.PhotoImage(icon)
        Label(self,image=self.photo).place(x=0,y=0)#to put  image in background
        
         #FIRST FRAME FOR HEADING
        fram=Frame(self,relief=RAISED)
        Label(fram,text="*FILL THE DETAILS :-",font='arialblack 15 bold',bg="yellow",fg='red').grid()
        fram.grid(column=1,row=0,padx=0,pady=50)
        
         #LABELS DECLARATION
        self.na=Label(self,text="NAME",font='arialblack 10 bold',bg="#88ffca")
        self.a=Label(self,text="AGE",font='arialblack 10 bold',bg="#88ffca")
        self.g=Label(self,text="GENDER",font='arialblack 10 bold',bg="#88ffca",)
        self.ph=Label(self,text="PHONE NO",font='arialblack 10 bold',bg="#88ffca",)
        self.ad=Label(self,text="ADDRESS",font='arialblack 10 bold',bg="#88ffca",)
        self.h=Label(self,text="HEIGHT (cm)",font='arialblack 10 bold',bg="#88ffca",)
        self.w=Label(self,text="WEIGHT (Kg)",font='arialblack 10 bold',bg="#88ffca",)
        self.bp=Label(self,text="BP LEVEL",font='arialblack 10 bold',bg="#88ffca",)
        self.ox=Label(self,text="OXYGEN LEVEL",font='arialblack 10 bold',bg="#88ffca",)

         #PACKING OF LABELS OF MAIN WINDOW
        self.na.grid(column=0,row=4,padx=100,pady=3)
        self.a.grid(column=0,row=5,padx=20,pady=4)
        self.g.grid(column=0,row=6,pady=4,padx=20)
        self.ph.grid(column=0,row=7,pady=4,padx=20)
        self.ad.grid(column=0,row=8,pady=4,padx=20)
        self.h.grid(column=0,row=9,pady=4,padx=20)
        self.w.grid(column=0,row=10,pady=4,padx=20)
        self.bp.grid(column=0,row=11,pady=4,padx=20)
        self.ox.grid(column=0,row=12,pady=4,padx=20)

         #DEFINE VARIABLE 
        nam=StringVar()
        ag=StringVar()
        phn=StringVar()
        add=StringVar()
        hei=IntVar()
        wei=IntVar()
        b_p=IntVar()
        oxy=IntVar()

          #ENTRY GRID FOR USER
        name=Entry(self,textvariable=nam)
        age=Entry(self,textvariable=ag)
        phone=Entry(self,textvariable=phn)
        addr=Entry(self,textvariable=add,)
        heig=Entry(self,textvariable=hei)
        weig=Entry(self,textvariable=wei)
        b1p=Entry(self,textvariable=b_p)
        oxyg=Entry(self,textvariable=oxy)
         
         #PACKING PF ENTRY 
        name.grid(column=1,row=4,padx=5,ipadx=40)
        age.grid(column=1,row=5,padx=5,ipadx=40)
         #gender selection
        combo=Combobox(self,state="readonly",justify='center')
        combo['values']=('SELECT','MALE','FEMALE','OTHER')
        combo.current(0)
        combo.grid(column=1,row=6,ipadx=10)
        phone.grid(column=1,row=7,padx=5,ipadx=40)
        addr.grid(column=1,row=8,padx=5,ipadx=40)
        heig.grid(column=1,row=9,padx=5,ipadx=40)
        weig.grid(column=1,row=10,padx=5,ipadx=40)
        b1p.grid(column=1,row=11,padx=5,ipadx=40)
        oxyg.grid(column=1,row=12,padx=5,ipadx=40)
        
         #VALIDATE COMMAND FOR NAME 
        validate_name=self.register(self.checkname)
        name.config(validate='key',validatecommand=(validate_name,'%P'))
        #VALIDATE COMMAND FOR phone no
        validate_phone=self.register(self.checkphone)
        phone.config(validate='key',validatecommand=(validate_phone,'%P'))
        #VALIDATE COMMAND FOR address
        ''' validate_adr=self.register(self.checkaddress)
        addr.config(validate='key',validatecommand=(validate_adr,'%P'))'''
        
          #END BUTTONS
        Button(self,text="CALCULATE",font="Times 13 bold",command=calculate,bg='white',fg='red',width=10).grid(column=1,row=14,pady=30)
        Button(self,text="RESET",font="bookmanoldstyle 13 bold",command=clear).grid(column=1,row=16)
        
        #LABEL FOR PROGRAMMERS NAME
        Label(self,text='MADE BY- SATYENDRA AND SARTHAK V',font="arialblak 9 bold ",bg='skyblue',fg='red').place(x=390,y=570)

        
        
        #FUNCTION TO CAKCULATE BMI
        def calculate_bmi(w,h):
            kg=int(w)
            m = int(h)/100
            bmi = kg/(m*m)
            bmi = round(bmi, 1)
            return (bmi)
        
        #TO CHECK BMI STATUS
        def bmi_index(bmi):
            if bmi < 18.5:
                return(" Is Underweight")
            elif (bmi > 18.5) and (bmi < 24.9):
              return(" Is Normal")
            elif (bmi > 24.9) and (bmi < 29.9):
                return(" Is Overweight")
            elif (bmi > 29.9):
                return(" Is Obesity") 
            else:
                return(" Something went wrong!!!") 
            
        #FUNCTION TO CHECK OXYGEN STATUS    
        def oxystatus(ox):
            oxy=int(ox)
            if oxy>=98 and oxy<=100:
                return('OXYGEN LEVEL is NORMAL')
            elif(oxy>=95 and oxy<=97):
                return('OXYGEN LEVEL is INSUFFIECIENT')
            elif(oxy>=90 and oxy<=94):
                return('OXYGEN LEVEL is DECREASED')
            elif(oxy<90):
                return('OXYGEN LEVEL is CRITICAL')
            elif(oxy<80):
                return('OXYGEN LEVEL is SEVERE HYPOXIA')
            elif(oxy<70):
                return('OXYGEN LEVEL is ACUTE DANGER TO LIFE')
            else:
                return('OOPS SOMETHING WENT WRONG!!!')
            
        #FUNCTION TO CHECK BP STATUS    
        def bpstatus(p):
            bp=int(p)
            if bp<80 :
                return('LOW BLOOD PRESSURE')
            elif bp>=80 and bp<120:
                return('NORMAL')
            elif bp>=120 and bp<=139 :
                return('PREHYPERTENSION')
            elif bp>=140 and bp<=159:
                return('HIGH BLOOD PRESSURE (STAGE-1)')
            elif bp>=160 and bp<180:
                return('HIGH BLOOD PRESSURE (STAGE-2)')
            elif bp>=180 :
                return('HIGH BLOOD PRESSURE (STAGE-3)')
            else:
                return('OOPS SOMETHING WENT WRONG !!!')
            
    #VALIDATE NAME ENTRY BY USER         
    def checkname(self,name):
            if name.isalpha(): 
                return True
            if name=='':
                return True
            #if len(str(name))>30:
                #messagebox.showerror('invalid','INVALID max character 30')
                #return False
            else :
                messagebox.showerror('invalid','Not allowed '+name[-1])   
           
    #VALIDATE PHONE ENTRY
    def checkphone(self,phone):
            Pattern = re.compile("(0|91)?[7-9][0-9]{9}")
            if phone.isdigit():
                return True
            if Pattern.match(phone):
                return True
            if len(str(phone))<11:
                return True
            else:
                messagebox.showerror('invalid','INVALID')
                return False
    
    #CREATION OF EXCEL FILE TO SAVE DATA 
    file = pathlib.Path('shinchan2.xlsx')
    if file.exists():
        pass
    else:
        file=Workbook()
        sheet=file.active
        sheet["A1"]="Name"
        sheet["B1"]="Age"
        sheet["C1"]="Gender"
        sheet["D1"]="Phone No"
        sheet["E1"]="Address"
        sheet["F1"]="Height"
        sheet["G1"]="Weight"
        sheet["H1"]="BMI"
        sheet["I1"]="BMI STATUS"
        sheet["J1"]="BP"
        sheet["K1"]="BP STATUS"
        sheet["L1"]="OXYGEN LEVEL"
        sheet["M1"]="OXYGEN STATUS"
        file.save('shinchan2.xlsx')    
              
#MAIN PROGRAM       
if __name__=="__main__":
    win=Lab()
    win.mainloop()
       
     
    
         